"""Generate `card_data.json` from `Slay the Spire Reference.xlsx`.

Turns the spreadsheet's `Cards` sheet into a static `CardId -> {type, rarity, cost,
text, ...}` table keyed by the engine's CardId enum names, so `card_text.py` and the
state encoder can look up by `card.id.name`.

Scope: only the sets the Ironclad-complete engine actually plays — Ironclad,
Colorless, Curses, Statuses.

Re-run after editing the sheet:  python game_data/card_data/build_card_data.py

The sheet encodes upgrades two ways, both handled here:
  * inline  "Deal 8 (10) damage."   -> base uses 8, upgraded uses 10
  * separate "Description (Upgraded)" column (when the change isn't just a number)
"""
import json
import re
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
XLSX = HERE.parent / "Slay the Spire Reference.xlsx"          # game_data/…
OUT = HERE / "card_data.json"
BINDINGS = REPO / "sts_lightspeed" / "bindings" / "slaythespire.cpp"

# (section label, first data row, last data row, CardColor) — row numbers are the
# 0-based indices into iter_rows() output, verified against the sheet layout.
SECTIONS = [
    ("Ironclad", 2, 76, "RED"),
    ("Colorless", 306, 356, "COLORLESS"),
    ("Curse", 358, 371, "CURSE"),
    ("Status", 373, 377, "COLORLESS"),
]

# Names whose CardId isn't the plain normalization of the display name.
NAME_TO_ID = {
    "Strike": "STRIKE_RED",   # Ironclad basic; engine color-suffixes it
    "Defend": "DEFEND_RED",
}


def normalize_id(name: str) -> str:
    """Display name -> CardId enum name (e.g. 'Blood for Blood' -> BLOOD_FOR_BLOOD)."""
    if name in NAME_TO_ID:
        return NAME_TO_ID[name]
    n = name.strip().upper().replace(".", "").replace("'", "")
    return re.sub(r"[^A-Z0-9]+", "_", n).strip("_")


def replace_energy_tokens(text: str) -> str:
    """Collapse runs of energy orbs ('[R] [R]') into readable text ('2 Energy')."""
    def repl(m):
        count = len(re.findall(r"\[[A-Z]\]", m.group(0)))
        return f"{count} Energy"
    return re.sub(r"\[[A-Z]\](?:\s*\[[A-Z]\])*", repl, text)


_INLINE = re.compile(r"(\d+)\s*\((\d+)\)")


def split_inline(text: str):
    """'Deal 8 (10) damage.' -> ('Deal 8 damage.', 'Deal 10 damage.')."""
    base = _INLINE.sub(lambda m: m.group(1), text)
    upg = _INLINE.sub(lambda m: m.group(2), text)
    return base, upg


def split_cost(cost) -> tuple[str, str | None]:
    """Cost cell -> (base_cost, upgraded_cost_or_None). Handles 'X', 'Unplayable', '1 (0)'."""
    s = str(cost).strip()
    m = _INLINE.fullmatch(s)
    if m:
        return m.group(1), m.group(2)
    return s, None


def load_enum_ids() -> set[str]:
    cpp = BINDINGS.read_text(encoding="utf-8")
    return set(re.findall(r'\.value\("([A-Z0-9_]+)",\s*CardId::', cpp))


def clean(text) -> str:
    return replace_energy_tokens(str(text).strip())


def main() -> None:
    enum_ids = load_enum_ids()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rows = list(wb["Cards"].iter_rows(values_only=True))

    cards: dict[str, dict] = {}
    problems: list[str] = []
    bracket_tokens: set[str] = set()

    for label, start, end, color in SECTIONS:
        for i in range(start, end + 1):
            row = (list(rows[i]) + [None] * 6)[:6]
            name, ctype, rarity, cost, desc, desc_up = row
            if not name or not str(name).strip():
                continue
            name = str(name).strip()

            cid = normalize_id(name)
            if cid not in enum_ids:
                problems.append(f"{label}: {name!r} -> {cid!r} not in CardId enum")

            bracket_tokens.update(re.findall(r"\[[^\]]+\]", str(desc)))

            if desc_up and str(desc_up).strip():
                # separate-column form: base as-is, upgraded from its own column
                text = clean(desc)
                text_up = clean(desc_up)
            else:
                base, upg = split_inline(clean(desc))
                text = base
                text_up = upg if upg != base else None

            cost_base, cost_up = split_cost(cost)

            cards[cid] = {
                "name": name,
                "type": str(ctype).strip().upper(),
                "rarity": str(rarity).strip().upper(),
                "color": color,
                "cost": cost_base,
                "cost_upgraded": cost_up,
                "text": text,
                "text_upgraded": text_up,
            }

    if problems:
        raise SystemExit("Unmapped cards:\n  " + "\n  ".join(problems))

    OUT.write_text(json.dumps(cards, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(cards)} cards -> {OUT.relative_to(REPO)}")
    print(f"Bracket tokens seen (should all be energy orbs): {sorted(bracket_tokens)}")
    by_color: dict[str, int] = {}
    for c in cards.values():
        by_color[c["color"]] = by_color.get(c["color"], 0) + 1
    print("By color:", by_color)


if __name__ == "__main__":
    main()
